import sys, re
from openpyxl import Workbook
from pprint import pprint
sys.path.insert(0, '../../modules')


import setting as p
import excel as img


def parcer():
    
    data = []
    for page in range(1,69):
        url = f'https://www.grupdirect.com/en?pag={page}#modulo-paginacion'
        print ('==================', url, f'==================={page}')
        soup = p.get_content_text(url)
        id = 1
        for i in soup.select('article.paginacion-ficha'):

            link = f"https://www.grupdirect.com/{i.select_one(' a.irAfichaPropiedad').get('href')}"[:-3]
            price = i.select_one('span.paginacion-ficha-tituloprecio').text
            print (link, id, price,)
            data.append(getData(price, link))
            # time.sleep(5)
            id+=1
            # pprint(data)
    makeExcel(data,'out/grupdirect.xlsx')

def getData(price,  url='https://www.grupdirect.com/ficha/flat/llucmajor/pueblo/3279/10685260'):
    lang = {'en': p.get_content_text(f'{url}/en/'),
            'es': p.get_content_text(f'{url}/es/'),
            'de': p.get_content_text(f'{url}/de/'),
            'ru': p.get_content_text(f'{url}/ru/')}
    title,  properties, features = {},{},{}
    price_rent, price_sale = '', ''

    def getProperties(soup):
        properties = {}
        for i in soup.select('ul.fichapropiedad-listadatos li'):
            properties.update({i.select_one('span.caracteristica').text:i.select_one('span.valor').text})
        return properties


    def  getFeatures(soup):
        features = []
        for i in soup.select('ul.fichapropiedad-listacalidades li'):
           features.append(i.text.replace('\xa0',''))
        return features
    

    for key, values in lang.items():
        try:
            title.update({key:values.select_one('div.fichapropiedad-tituloprincipal').text.strip()})
        except:
            title.update({key:''})
        try:
            properties.update({key:getProperties(values)})
        except:
            properties.update({key:''})
        try:
            features.update({key:getFeatures(values)})
        except:
            features.update({key:''})
    
    images = []
    try:
        for i in lang['en'].select('div.visorficha-miniaturas li'):
            images.append(f"{i.get('cargafoto')[:-5]}{i.get('cargafoto')[-4:]}")
    except:
        images.append('')

    if 'For sale' in price:
        price_sale = price
    if 'Rental' in price:
        price_rent = price
    return{
        'title':title,
        'location':getProperties(lang['en'])['Area / City'],
        'images':images,
        'properties':properties,
        'features':features,
        'ref' : getProperties(lang['en'])['Reference'],
        'price_sale': ''.join(re.findall(r'[0-9]+', price_sale)),
        'price_rent': ''.join(re.findall(r'[0-9]+', price_rent)),
        'url':url
    }
  

def makeExcel(data, filename):

    def getProperties(data):
        prop = []
        for key, values in data.items():
            prop.append(f'{key} : {values}')
        return ', '.join(prop)


    wb = Workbook()
    header = ['ref',
              'price_sale',
              'price_rent',
              'title_en', 
              'title_es',
              'title_de',
              'title_ru',
              'location',
              'features_en',
              'features_es',
              'features_de',
              'features_ru',
              'properties_en',
              'properties_es',
              'properties_de',
              'properties_ru',
              'url',
              'images']
    target = wb.active
    row = 2
    col = 1
    for i in header:
        target.cell(1, col).value = i
        col += 1
    for i in data:
        target.cell(row, 1).value = i['ref']
        target.cell(row, 2).value = i['price_sale']
        target.cell(row, 3).value = i['price_rent']
        target.cell(row, 4).value = i['title']['en']
        target.cell(row, 5).value = i['title']['es']
        target.cell(row, 6).value = i['title']['de']
        target.cell(row, 7).value = i['title']['ru']
        target.cell(row, 8).value = i['location']
        target.cell(row, 9).value = ', '.join(i['features']['en'])
        target.cell(row, 10).value = ', '.join(i['features']['es'])
        target.cell(row, 11).value = ', '.join(i['features']['de'])
        target.cell(row, 12).value = ', '.join(i['features']['ru'])
        target.cell(row, 13).value = getProperties(i['properties']['en'])
        target.cell(row, 14).value = getProperties(i['properties']['es'])
        target.cell(row, 15).value = getProperties(i['properties']['de'])
        target.cell(row, 16).value = getProperties(i['properties']['ru'])
        target.cell(row, 17).value = i['url']
        target.cell(row, 18).value = ','.join(i['images'])

       
        row += 1

    wb.save(filename)

if __name__ == "__main__":
    start = int(input('start='))
    stop = int(input('stop='))
    row = int(input('row='))
    img.dImages('grupdirect.xlsx', start=start, row=row, stop=stop)
    # parcer()
    # pprint(getData('10000', 'Sale'))